import pandas as pd
import os
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font

try:
    os.system("rm -rf user_data")
except:
    pass

# Load the CSV files
submissions = pd.read_csv("submissions.csv")
image_labels = pd.read_csv("image_labels.csv")

# Ensure column names are stripped of whitespace (if any)
submissions.columns = submissions.columns.str.strip()
image_labels.columns = image_labels.columns.str.strip()

# Get unique users
unique_users = submissions["User Name"].unique()

# Base directory for user data
base_dir = "user_data"
os.makedirs(base_dir, exist_ok=True)

# Root directory for images
image_root_dir = "static/fullres"

# Mapping for predicted_label_final
label_mapping = {
    "Definitely Real": "Real",
    "Probably Real": "Real",
    "Unsure": "Unsure",
    "Definitely Fake": "Fake",
    "Probably Fake": "Fake"
}

# Process each user
for user in unique_users:
    # Create a directory for each user
    user_dir = os.path.join(base_dir, user)
    os.makedirs(user_dir, exist_ok=True)
    
    # Create subdirectories for TP, TN, FP, FN, and Unsure
    tp_dir = os.path.join(user_dir, "TP")
    tn_dir = os.path.join(user_dir, "TN")
    fp_dir = os.path.join(user_dir, "FP")
    fn_dir = os.path.join(user_dir, "FN")
    unsure_dir = os.path.join(user_dir, "Unsure")
    os.makedirs(tp_dir, exist_ok=True)
    os.makedirs(tn_dir, exist_ok=True)
    os.makedirs(fp_dir, exist_ok=True)
    os.makedirs(fn_dir, exist_ok=True)
    os.makedirs(unsure_dir, exist_ok=True)

    # Filter data for the user
    user_data = submissions[submissions["User Name"] == user]

    # Merge user data with image labels based on filename
    merged_data = user_data.merge(image_labels, left_on="File", right_on="Filename", how="inner")

    # Select relevant columns
    final_data = merged_data[["File", "Label", "RealFakeValueSlider"]]
    final_data.columns = ["file_name", "original_label", "predicted_label"]

    # Apply mapping to create predicted_label_final column
    final_data.loc[:, "predicted_label_final"] = final_data["predicted_label"].map(label_mapping)

    # Compute TP, TN, FP, FN columns
    final_data.loc[:, "TP"] = ((final_data["original_label"] == "Fake") & (final_data["predicted_label_final"] == "Fake")) * 1
    final_data.loc[:, "TN"] = ((final_data["original_label"] == "Real") & (final_data["predicted_label_final"] == "Real")) * 1
    final_data.loc[:, "FN"] = ((final_data["original_label"] == "Fake") & (final_data["predicted_label_final"] == "Real")) * 1
    final_data.loc[:, "FP"] = ((final_data["original_label"] == "Real") & (final_data["predicted_label_final"] == "Fake")) * 1
    final_data.loc[:, "Unsure"] = (final_data["predicted_label_final"] == "Unsure") * 1

    # Copy images into respective folders
    for _, row in final_data.iterrows():
        source_path = os.path.join(image_root_dir, row["file_name"])
        if os.path.exists(source_path):
            if row["TP"] == 1:
                shutil.copy(source_path, tp_dir)
            elif row["TN"] == 1:
                shutil.copy(source_path, tn_dir)
            elif row["FP"] == 1:
                shutil.copy(source_path, fp_dir)
            elif row["FN"] == 1:
                shutil.copy(source_path, fn_dir)
            elif row["Unsure"] == 1:
                shutil.copy(source_path, unsure_dir)
    
    # Compute total values
    total_TP = final_data["TP"].sum()
    total_TN = final_data["TN"].sum()
    total_FN = final_data["FN"].sum()
    total_FP = final_data["FP"].sum()
    total_Unsure = final_data["Unsure"].sum()
    just_total = total_TP + total_TN + total_FN + total_FP + total_Unsure
    just_total_excluding_unsure = total_TP + total_TN + total_FN + total_FP

    # Compute precision, recall, and accuracy
    precision = total_TP / (total_TP + total_FP) if (total_TP + total_FP) > 0 else 0
    recall = total_TP / (total_TP + total_FN) if (total_TP + total_FN) > 0 else 0
    accuracy = (total_TP + total_TN) / (total_TP + total_TN + total_FP + total_FN) if (total_TP + total_TN + total_FP + total_FN) > 0 else 0

    # Append summary statistics with descriptions
    summary_data = pd.DataFrame({
        "Metric": ["Total Attempted", "Total Excluding Unsure", "Total TP", "Total TN", "Total FN", "Total FP", "Total Unsure", "Precision", "Recall", "Accuracy"],
        "Value": [just_total, just_total_excluding_unsure, total_TP, total_TN, total_FN, total_FP, total_Unsure, precision, recall, accuracy],
        "Description": [
            "Sum total of all images attempted by the user including 'Unsure' predictions.",
            "Sum total of all images attempted by the user excluding 'Unsure' predictions.",
            "True Positive (TP): The original image was 'Fake' and was correctly predicted as 'Fake.'",
            "True Negative (TN): The original image was 'Real' and was correctly predicted as 'Real.'",
            "False Negative (FN): The original image was 'Fake' but was incorrectly predicted as 'Real.'",
            "False Positive (FP): The original image was 'Real' but was incorrectly predicted as 'Fake.'",
            "Unsure: The image was predicted as 'Unsure.'",
            "Precision: Out of all the images predicted as 'Fake,' Precision tells us how many were actually 'Fake.', calculated as TP / (TP + FP).",
            "Recall: Out of all the images that were truly 'Fake,' Recall tells us how many were correctly identified as 'Fake.' calculated as TP / (TP + FN).",
            "Accuracy: The overall correctness of the user, calculated as (TP + TN) / (TP + TN + FP + FN)."
        ]
    })
    
    # Save to an Excel file in the user's directory
    user_filename = os.path.join(user_dir, f"{user}.xlsx")
    with pd.ExcelWriter(user_filename, engine="openpyxl") as writer:
        final_data.to_excel(writer, index=False, sheet_name="Results")
        summary_data.to_excel(writer, index=False, sheet_name="Summary")

    # Load workbook and apply styles
    wb = load_workbook(user_filename)
    ws = wb["Summary"]
    for cell in ws["A"]:
        cell.font = Font(bold=True)
    wb.save(user_filename)

print("User-wise Excel files with summary statistics and image categorization (including Unsure images) have been generated successfully in their respective folders.")
